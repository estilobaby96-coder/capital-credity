"""Motor de Exportação — PDF (reportlab) e Excel (openpyxl) com logo da empresa."""

import os
from datetime import date, datetime
from typing import List, Dict, Any

# ── PDF ──
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Excel ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


# Caminho absoluto da logo
from utils.resource_path import get_resource_path
LOGO_PATH = get_resource_path(os.path.join("assets", "logo", "logo.png"))

# Cores da identidade visual
VERDE_PRIMARIO = colors.HexColor("#00C853")
VERDE_ESCURO = colors.HexColor("#1B5E20")
CINZA_CLARO = colors.HexColor("#F5F5F5")
CINZA_ESCURO = colors.HexColor("#333333")


class ExportService:
    """Gera arquivos PDF e Excel profissionais com cabeçalho, logo e tabelas formatadas."""

    # ================================================================== #
    #  PDF                                                                 #
    # ================================================================== #
    def exportar_pdf(self, filepath: str, titulo: str, colunas: List[str],
                     dados: List[List[str]], subtitulo: str = "") -> str:
        """
        Gera um PDF com logo, título e tabela formatada.
        
        Args:
            filepath: Caminho completo de saída (.pdf)
            titulo: Título do relatório
            colunas: Lista com nomes das colunas
            dados: Lista de linhas (cada linha é uma lista de strings)
            subtitulo: Texto auxiliar (ex: período, nome do cliente)
        Returns:
            filepath do arquivo gerado
        """
        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            topMargin=15 * mm,
            bottomMargin=15 * mm,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # ── Cabeçalho com Logo ──
        header_data = []
        logo_cell = ""
        if os.path.exists(LOGO_PATH):
            logo_cell = RLImage(LOGO_PATH, width=50, height=50)

        titulo_style = ParagraphStyle(
            "TituloRelatorio", parent=styles["Title"],
            fontSize=18, textColor=CINZA_ESCURO, spaceAfter=2
        )
        sub_style = ParagraphStyle(
            "SubRelatorio", parent=styles["Normal"],
            fontSize=10, textColor=colors.gray
        )
        data_style = ParagraphStyle(
            "DataRelatorio", parent=styles["Normal"],
            fontSize=9, textColor=colors.gray, alignment=TA_RIGHT
        )

        titulo_p = Paragraph(titulo, titulo_style)
        sub_p = Paragraph(subtitulo, sub_style) if subtitulo else Paragraph("", sub_style)
        data_p = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", data_style)

        header_table = Table(
            [[logo_cell, [titulo_p, sub_p], data_p]],
            colWidths=[60, None, 150]
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 8 * mm))

        # ── Linha divisória ──
        div = Table([[""]],
                    colWidths=[doc.width],
                    rowHeights=[2])
        div.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), VERDE_PRIMARIO),
        ]))
        elements.append(div)
        elements.append(Spacer(1, 5 * mm))

        # ── Tabela de Dados ──
        cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)
        header_cell_style = ParagraphStyle(
            "HeaderCell", parent=styles["Normal"],
            fontSize=9, textColor=colors.white, leading=11
        )

        # Converter todos os dados para Paragraphs (para word-wrap)
        table_data = [[Paragraph(str(c), header_cell_style) for c in colunas]]
        for row in dados:
            table_data.append([Paragraph(str(cell), cell_style) for cell in row])

        num_cols = len(colunas)
        col_width = doc.width / num_cols

        data_table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)

        # Estilo da tabela
        style_cmds = [
            # Cabeçalho
            ("BACKGROUND", (0, 0), (-1, 0), VERDE_ESCURO),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        # Zebra striping
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), CINZA_CLARO))

        data_table.setStyle(TableStyle(style_cmds))
        elements.append(data_table)

        # ── Rodapé ──
        elements.append(Spacer(1, 10 * mm))
        footer_style = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
        elements.append(Paragraph("Capital Credity", footer_style))

        doc.build(elements)
        return filepath

    # ================================================================== #
    #  EXCEL                                                               #
    # ================================================================== #
    def exportar_excel(self, filepath: str, titulo: str, colunas: List[str],
                       dados: List[List[str]], subtitulo: str = "") -> str:
        """
        Gera um Excel (.xlsx) com logo embutida, título e tabela formatada.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # ── Estilos ──
        verde_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="333333")
        sub_font = Font(name="Calibri", size=10, color="888888")
        data_font = Font(name="Calibri", size=10, color="333333")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # ── Logo ──
        current_row = 1
        if os.path.exists(LOGO_PATH):
            try:
                img = XLImage(LOGO_PATH)
                img.width = 80
                img.height = 80
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 65
                current_row = 2
            except Exception:
                pass  # Logo indisponível, segue sem

        # ── Título e Subtítulo ──
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(colunas))
        cell_titulo = ws.cell(row=current_row, column=2, value=titulo)
        cell_titulo.font = title_font

        current_row += 1
        if subtitulo:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(colunas))
            cell_sub = ws.cell(row=current_row, column=2, value=subtitulo)
            cell_sub.font = sub_font
            current_row += 1

        # Data de geração
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(colunas))
        cell_data = ws.cell(row=current_row, column=2, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        cell_data.font = sub_font
        current_row += 2  # Linha em branco

        # ── Cabeçalho da Tabela ──
        header_row = current_row
        for col_idx, col_name in enumerate(colunas, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = verde_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 28
        current_row += 1

        # ── Dados ──
        zebra_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for row_idx, row_data in enumerate(dados):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
            current_row += 1

        # ── Auto-ajustar largura das colunas ──
        for col_idx in range(1, len(colunas) + 1):
            max_len = len(str(colunas[col_idx - 1]))
            for row in dados:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            adjusted = min(max_len + 4, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        wb.save(filepath)
        return filepath
